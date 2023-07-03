import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar


# cores
cor1 = "#89004f"  # magenta
cor2 = "#ffe0ff"  # rosa clarinho
cor3 = "#5ff4ab"  # verde
cor4 = "#f70071"  # vermelha
cor5 = "#c33b80"  # rosa escuro
cor6 = "#6bb5ff"  # azul~
cor7 = "#ff69b4"  # rosa
cor8 = "#ffa8d9"  # rosa claro


# função salvar valores
def salvar_valor():
    valor_dia = float(entry_valor.get())
    valores.append(valor_dia)
    total_valores = sum(valores)
    entry_valor.delete(0,tk.END)

    label_status.config(text="Valor salvo com sucesso!", foreground=cor3)
    label_total.config(text=f"Total na poupança: R${total_valores:.2f}")

    # adicionando valor em uma nova linha na planilha
    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1,value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha, column=2, value=valor_dia)


# função para plotar gráficos

def plotar_grafico():
    global canvas

    #obtendo as datas e os valores
    datas = [cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, "%d-%m-%y").date() for cell in sheet['A'][1:]]
    valores = [cell.value for cell in sheet['B'][1:]]

    # agrupando valores por mês
    dados_mensais = {}
    for data, valor in zip(datas, valores):
        mes_ano = data.strftime("%m-%Y")
        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]
            
    fig = plt.Figure(figsize=(12, 6), dpi=80)
    ax_barras = fig.add_subplot(121)
    ax_pie = fig.add_subplot(122)

    barras = ax_barras.bar(range(len(dados_mensais)), [sum(valores) for valores in dados_mensais.values()])

    for i, barra in enumerate(barras):
        altura = barra.get_height()
        ax_barras.text(barra.get_x() + barra.get_width() / 2, altura, f'R${altura:.2f}', ha='center', va='bottom')

    nomes_meses = []
    for mes_ano in dados_mensais.keys():
        mes, ano = mes_ano.split('-')
        nome_mes = calendar.month_name[int(mes)]
        nomes_meses.append(f'{nome_mes}-{ano}')

    ax_barras.set_xticks(range(len(dados_mensais)))
    ax_barras.set_xticklabels(nomes_meses, ha='right')
    
    ax_barras.spines['top'].set_visible(False)
    ax_barras.spines['right'].set_visible(False)
    ax_barras.spines['bottom'].set_visible(False)
    ax_barras.spines['left'].set_visible(False)
    
    ax_barras.set_title('Economia por Mês')
    ax_barras.title.set_position([.5, 8.05])
    ax_barras.set_xlabel('Mês')
    ax_barras.set_ylabel('Valor Economizado')


    # analisando semanalmente
    data_inicial = min(datas)
    data_final = max(datas)

    diferenca = (data_final - data_inicial).days
    semanas = diferenca//7
    
    labels = [f'{i+1}ª Semana' for i in range(semanas)]
    valores_semana = []
    for i in range(semanas):
        data_inicio = data_inicial + timedelta(weeks=i)
        data_fim = data_inicio + timedelta(weeks=1)
        valores_semana.append(sum(valor for data, valor in zip(datas, valores) if data_inicio <= data < data_fim))

    pie = ax_pie.pie(valores_semana, labels=labels, autopct='%1.1f%%', startangle=90)
    ax_pie.set_title('Economia por Semana')

    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.get_tk_widget().grid(row=5, column=0, columnspan=2, padx=10, pady=10)

    fig.tight_layout()


    # exibindo o gráfico no Tkinter
    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)



janela = tk.Tk()
janela.title = ('App de Poupança Pessoal')
janela.geometry("700x500")
janela.configure(bg=cor1) 


style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel", background=cor1, foreground= cor2, font=("Arial, 12"))
style.configure("TEntry", fieldbackground= cor2, font=("Arial, 12"))
style.configure("TButton", background=cor5, foreground= cor2, font=("Arial, 12"))

label_instrucao = ttk.Label(janela, text="Insira o valor diário")
label_status = ttk.Label(janela, text="aaaa", foreground=cor8)
label_total = ttk.Label(janela, text="sssss", font=("Arial", 14, "bold"))
entry_valor = ttk.Entry(janela)
button_salvar = ttk.Button(janela, text="Salvar", command=salvar_valor)


# posicionando os elementos
label_instrucao.pack(pady=10)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack()
label_total.pack(pady=10)


# carrengado a planilha do excel
try:
    workbook = load_workbook("valores_diarios.xlsx")
except FileNotFoundError:
    workbook=Workbook()


# Selecionando a primeira planilha
sheet = workbook.active

# verificando se a planilha já possui valores salvos

if sheet.max_row == 0:
    sheet.cell(row=1, column=1, value="Data")
    sheet.cell(row=1, column=2, value="Valor diário")

# obtendo a lista de valores já salvos
valores = [cell.value for cell in sheet['B'][1:]]

# exibindo o total economizado
label_total.config(text=f'Total economizado: R${sum(valores):.2f}')

# plotando o gráfico inicial
plotar_grafico()

janela.mainloop()


# salvando a planilha com os valores atualizados
workbook.save("valores_diarios.xlsx")